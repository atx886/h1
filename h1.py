import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import date, timedelta
from decimal import *
import os
import notify
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.header import Header
import time
from bs4 import BeautifulSoup

path = os.getcwd()


def 自动下载文件(zh, mm):
    last_time = str(date.today() + timedelta(days=-1))
    now_time = str(date.today())

    headers = {
        'Connection': 'keep-alive',
        'Cache-Control': 'max-age=0',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.80 Mobile Safari/537.36 Edg/98.0.1108.43',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6'
    }
    url_1 = 'http://life.wuzhouhongye.com/subsidiary/cabinet_order_goods?pay_status=1&created_at%5Bstart%5D=' + last_time + '+00%3A00%3A00&created_at%5Bend%5D=' + now_time + '+00%3A00%3A00&_pjax=%23pjax-container&_export_=all'
    url_2 = 'http://life.wuzhouhongye.com/subsidiary/cabinet_order_goods?pay_status=0&created_at%5Bstart%5D=' + last_time + '+00%3A00%3A00&created_at%5Bend%5D=' + now_time + '+00%3A00%3A00&_pjax=%23pjax-container&_export_=all'
    login = 'http://life.wuzhouhongye.com/subsidiary/auth/login'
    session = requests.session()
    t = session.get(url=login, headers=headers)
    soup = BeautifulSoup(t.text, 'html.parser')
    token = soup.input['value']
    print(token)
    datas = {
        "_token": token,
        "username": zh,
        "password": mm,
        'remember': "1"
    }
    print(datas)

    session.post(url=login, data=datas, headers=headers)
    r = session.get(url_1, allow_redirects=False)
    # r = requests.get(url_1, headers=headers)
    print(r.status_code)
    with open("1.xlsx", "wb") as f:
        f.write(r.content)
    # r = requests.get(url_2, headers=headers)
    r = session.get(url_2, headers=headers)
    with open("2.xlsx", "wb") as f:
        f.write(r.content)


def 删除文件():
    path1 = path + "/1.xlsx"
    path2 = path + "/2.xlsx"
    path3 = path + "/" + (str((date.today() + timedelta(days=-2))) + '未付款.xlsx')
    path4 = path + "/" + (str((date.today() + timedelta(days=-2))) + '.xlsx')
    if os.path.exists(path1):
        os.remove(path1)
    if os.path.exists(path2):
        os.remove(path2)
    if os.path.exists(path3):
        os.remove(path3)
    if os.path.exists(path4):
        os.remove(path4)


def 发送邮件(fasong, jieshou, miyao):
    # 1. 连接邮箱服务器
    con = smtplib.SMTP_SSL('smtp.163.com', 465)
    youxiangjieshou = jieshou
    # 2. 登录邮箱
    con.login(fasong, miyao)
    # 2. 准备数据
    # 创建邮件对象
    msg = MIMEMultipart()
    # 设置邮件主题
    subject = Header('文件附件发送', 'utf-8').encode()
    msg['Subject'] = subject
    # 设置邮件发送者
    msg['From'] = 'hhh'
    # 设置邮件接受者
    msg['To'] = youxiangjieshou
    # 添加文件附件
    succesful = str((date.today() + timedelta(days=-1))) + '.xlsx'
    faile = str((date.today() + timedelta(days=-1))) + '未付款.xlsx'
    file1 = MIMEText(open(succesful, 'rb').read(), 'base64', 'utf-8')
    file1["Content-Disposition"] = 'attachment; filename="' + succesful + '"'
    msg.attach(file1)
    # 中文时使用下面方法
    if os.path.exists(path + "/" + (str((date.today() + timedelta(days=-1))) + '未付款.xlsx')):
        file2 = MIMEText(open(faile, 'rb').read(), 'base64', 'utf-8')
        file2.add_header("Content-Disposition", 'attachment', filename=('gbk', '', faile))
        msg.attach(file2)
    # 3.发送邮件
    con.sendmail('sch1532694569s@163.com', youxiangjieshou, msg.as_string())
    print("发送成功")
    con.quit()


def 支付成功():
    wb = load_workbook('1.xlsx')
    sheet = wb.active

    max_row = sheet.max_row
    max_col = sheet.max_column
    k = max_row
    row_max = 'a' + str(max_row)
    # print(row_max)
    a1 = 1
    while a1 <= k:
        sheet['a' + str(a1)] = a1 - 1
        # print(sheet['a'+str(a1)].value)
        a1 += 1
    sheet['a1'] = "序号"
    # print(sheet['a1'])
    # sheet['a2'] = 1
    # wb.save('1.xlsx')
    s = ['序号', '柜子编号', '商品名字', '支付价格', '支付积分', '数量', '支付时间', '下单时间', '订单状态', '支付方式', '柜子小区']
    num = [i for i in range(max_col + 1)]
    a = 1

    # 获取行列 前行 后列
    while a <= max_col:
        for t in s:
            if sheet.cell(1, a).value == t:
                print(sheet.cell(1, a).value)
                num[a] = 0
        a += 1
    print(num)
    # print('成功获取项目表')
    i = 0
    for a in num:
        if a != 0:
            sheet.delete_cols(a - i)
            i += 1

    # 设置单元格的大小
    # sheet.row_dimensions['c'].width = 20.0
    sheet.column_dimensions['b'].width = 10.0
    sheet.column_dimensions['c'].width = 20.0
    sheet.column_dimensions['g'].width = 20.0
    sheet.column_dimensions['h'].width = 20.0
    sheet.column_dimensions['k'].width = 20.0

    sum1 = Decimal(sheet['d2'].value)
    sum2 = Decimal(sheet['e2'].value)
    for a in range(max_row - 2):
        sum1 += Decimal(sheet['d' + str(a + 3)].value)
    for a in range(max_row - 2):
        sum2 += Decimal(sheet['e' + str(a + 3)].value)
    sheet['d' + str(max_row + 1)] = sum1
    sheet['e' + str(max_row + 1)] = sum2

    # 添加一行并且合并 设置第一行为时间
    sheet.insert_rows(1)
    sheet1 = wb.sheetnames
    ws = wb[sheet1[0]]
    ws.merge_cells("A1:K1")
    sheet['a1'] = (date.today() + timedelta(days=-1)).strftime('%Y/%m/%d')

    # 设置居中
    for row in sheet:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(str((date.today() + timedelta(days=-1))) + '.xlsx')


def 未付款():
    wb = load_workbook('2.xlsx')
    sheet = wb.active
    if sheet['a1'].value == "ID":
        print("有东西")
    else:
        return
    max_row = sheet.max_row
    max_col = sheet.max_column
    k = max_row
    row_max = 'a' + str(max_row)
    # print(row_max)
    a1 = 1
    while a1 <= k:
        sheet['a' + str(a1)] = a1 - 1
        # print(sheet['a'+str(a1)].value)
        a1 += 1
    sheet['a1'] = "序号"
    print(sheet['a1'])
    # sheet['a2'] = 1
    # wb.save('1.xlsx')
    s = ['序号', '柜子编号', '商品名字', '支付价格', '支付积分', '数量', '支付时间', '下单时间', '订单状态', '支付方式', '柜子小区']
    num = [i for i in range(max_col + 1)]
    a = 1

    # 获取行列 前行 后列
    while a <= max_col:
        for t in s:
            if sheet.cell(1, a).value == t:
                print(sheet.cell(1, a).value)
                num[a] = 0
        a += 1
    print(num)
    i = 0
    for a in num:
        if a != 0:
            sheet.delete_cols(a - i)
            i += 1

    # 设置单元格的大小
    # sheet.row_dimensions['c'].width = 20.0
    sheet.column_dimensions['b'].width = 10.0
    sheet.column_dimensions['c'].width = 20.0
    sheet.column_dimensions['g'].width = 20.0
    sheet.column_dimensions['h'].width = 20.0
    sheet.column_dimensions['k'].width = 20.0

    sum1 = Decimal(sheet['d2'].value)
    sum2 = Decimal(sheet['e2'].value)
    for a in range(max_row - 2):
        sum1 += Decimal(sheet['d' + str(a + 3)].value)
    for a in range(max_row - 2):
        sum2 += Decimal(sheet['e' + str(a + 3)].value)
    sheet['d' + str(max_row + 1)] = sum1
    sheet['e' + str(max_row + 1)] = sum2

    # 添加一行并且合并 设置第一行为时间
    sheet.insert_rows(1)
    sheet1 = wb.sheetnames
    ws = wb[sheet1[0]]
    ws.merge_cells("A1:K1")
    sheet['a1'] = (date.today() + timedelta(days=-1)).strftime('%Y/%m/%d')

    # 设置居中
    for row in sheet:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(str((date.today() + timedelta(days=-1))) + '未付款.xlsx')


def init_main(zh, mm, fasong, jieshou, miyao):
    删除文件()
    time.sleep(1)
    自动下载文件(zh, mm)
    支付成功()
    未付款()
    time.sleep(2)

    发送邮件(fasong, jieshou, miyao)
    time.sleep(2)


zh = os.environ['zh']
mm = os.environ['mm']
miyao = os.environ['miyao']
fasong = os.environ['fasong']
jieshou = os.environ['jieshou']

init_main(zh, mm, fasong, jieshou, miyao)
notify.send('发送', 'ok')
